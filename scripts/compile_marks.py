"""Compile the finished listening audit into manifests + summary stats.

Reads the verdicts from reports/listening_audit.xlsx, merges full sample
metadata, and writes:
  reports/rejected_samples.csv   REJECT rows with metadata (+ reason/notes)
  reports/kept_samples.csv       KEEP rows
  reports/borderline_samples.csv BORDERLINE rows

Read-only on the workbook.
"""

import sys
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "reports" / "listening_audit.xlsx"
MANIFEST = ROOT / "data" / "manifests" / "sample_200.csv"
REPORTS = ROOT / "reports"

VERDICTS = ["KEEP", "BORDERLINE", "REJECT"]
TOTAL_HOURS = 1142.0


def read_marks() -> pd.DataFrame:
    wb = load_workbook(XLSX, read_only=True)
    ws = wb["AUDIT"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        order, wav, stratum, dur, auto_flags, text, verdict, reason, note = row[:9]
        if order is None:
            break
        rows.append({
            "listen_order": order, "wav_file": wav, "auto_flags": auto_flags or "",
            "verdict": str(verdict).strip().upper() if verdict else "",
            "reason": str(reason).strip().lower() if reason else "",
            "notes": note or "",
        })
    wb.close()
    return pd.DataFrame(rows)


def main() -> None:
    marks = read_marks()
    m = pd.read_csv(MANIFEST)
    df = marks.merge(m, on="wav_file", how="left")

    judged = marks["verdict"].isin(VERDICTS).sum()
    counts = marks["verdict"].value_counts()
    print(f"progress: {judged}/{len(marks)} judged\n")
    for v in VERDICTS:
        n = int(counts.get(v, 0))
        print(f"  {v:<11} {n:>4}  ({100 * n / judged:.1f}%)")

    keep_frac = counts.get("KEEP", 0) / judged
    rej_frac = counts.get("REJECT", 0) / judged
    print(f"\ncorpus projection ({TOTAL_HOURS:.0f} h):")
    print(f"  keep       ~{keep_frac * TOTAL_HOURS:.0f} h")
    print(f"  borderline ~{counts.get('BORDERLINE', 0) / judged * TOTAL_HOURS:.0f} h")
    print(f"  reject     ~{rej_frac * TOTAL_HOURS:.0f} h")

    print("\nby stratum:")
    g = df.groupby("stratum")["verdict"].value_counts().unstack(fill_value=0)
    for v in VERDICTS:
        if v not in g:
            g[v] = 0
    print(g[VERDICTS].to_string())

    flagged = df[df["auto_flags"] != ""]
    clean = df[df["auto_flags"] == ""]
    print("\nDSP auto-flags vs ears:")
    for name, part in [("flagged", flagged), ("unflagged", clean)]:
        n = len(part)
        v = part["verdict"].value_counts()
        print(f"  {name:<9} n={n:>3}  keep={v.get('KEEP',0)}  borderline={v.get('BORDERLINE',0)}  reject={v.get('REJECT',0)}")

    reasons = df[df["reason"] != ""]["reason"].value_counts()
    if len(reasons):
        print("\nreject/borderline reasons:")
        for r, n in reasons.items():
            print(f"  {r:<22} {n}")
    else:
        print("\n(reason column empty -- no reason breakdown available)")

    # ---- manifests ----
    meta_cols = ["listen_order", "wav_file", "stratum", "shard", "row_in_shard",
                 "duration_ss", "mean_speech_prob", "dbfs", "corrected_text",
                 "auto_flags", "reason", "notes"]
    meta_cols = [c for c in meta_cols if c in df.columns]
    for verdict, fname in [("REJECT", "rejected_samples.csv"),
                           ("KEEP", "kept_samples.csv"),
                           ("BORDERLINE", "borderline_samples.csv")]:
        part = df[df["verdict"] == verdict][meta_cols].sort_values("listen_order")
        part.to_csv(REPORTS / fname, index=False)
        print(f"[done] {len(part):>3} rows -> reports/{fname}")


if __name__ == "__main__":
    sys.exit(main())
