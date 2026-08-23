"""Build the Excel listening-audit workbook.

Output: reports/listening_audit.xlsx

Sheets:
  INSTRUCTIONS  how to mark, the 9 listening criteria, priority order, reasons
  AUDIT         all 200 clips, dropdowns for VERDICT + REASON, wav hyperlinks,
                colour-coded verdicts. Fill this sheet while listening.
  GOOD          live view of KEEP rows        (Excel 365/2021 FILTER formula)
  BORDERLINE    live view of BORDERLINE rows
  BAD           live view of REJECT rows

On Excel versions without dynamic arrays the GOOD/BORDERLINE/BAD sheets will
show #NAME? -- filter the AUDIT sheet by the VERDICT column instead.
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
SHEET_CSV = ROOT / "reports" / "listening_sheet.csv"
OUT = ROOT / "reports" / "listening_audit.xlsx"

VERDICTS = ["KEEP", "BORDERLINE", "REJECT"]
REASONS = [
    "cut_start", "cut_end", "multiple_speakers", "noise", "music",
    "clipping", "transcript_mismatch", "non_ghanaian", "unnatural_delivery",
    "other",
]

INSTRUCTIONS = """GHANA TTS LISTENING AUDIT
========================

Goal: decide whether each clip is good training material for a
Ghanaian-English TTS model. Don't overthink it.

WORKFLOW
--------
1. Work down the AUDIT sheet in order (flagged suspects first).
2. Play the clip (click the wav_file link, or use the HTML player).
3. Set VERDICT (dropdown):   KEEP / BORDERLINE / REJECT
4. If not KEEP, set one REASON (dropdown).
5. Add NOTES only when something needs explaining.

VERDICTS
--------
KEEP        good training clip.
BORDERLINE  usable, but something questionable.
REJECT      clearly should not go into the training set.

REASON CODES (pick one)
-----------------------
cut_start              clip starts mid-word / mid-sentence
cut_end                clip ends mid-word / mid-sentence (15 s ceiling!)
multiple_speakers      second voice, interruption, interviewer+interviewee
noise                  heavy static / distortion / loud background / echo
music                  bed, jingle, song under the speech
clipping               audible digital distortion at peaks
transcript_mismatch    words/names/numbers differ from what is spoken
non_ghanaian           not genuinely Ghanaian English
unnatural_delivery     shouting, chanting, singing, crowd responses
other                  anything else (explain in NOTES)

WHAT TO LISTEN FOR (9 checks)
-----------------------------
1. One speaker?            REJECT overlap, interruptions, second voices.
2. Ghanaian English?       pronunciation/rhythm/accent genuinely Ghanaian.
3. Clean audio?            light background noise is OK; drowning noise isn't.
4. Clean start?            first second: no missing word-half, no mid-sentence.
5. Clean ending?           THE BIG ONE. 15 s hard ceiling -> many cuts.
6. Transcript match?       read corrected_text AFTER listening. Wrong words/
                           names/numbers matter; punctuation doesn't.
7. Music?                  speech over noticeable music -> lean reject.
8. Natural delivery?       shouting/chanting/singing/long pauses -> not useful.
9. Standalone sentence?    "...and this is why we believe..." is weak training
                           material; complete utterances are best.

PRIORITY ORDER (what to reject on first)
----------------------------------------
1. multiple speakers        4. heavy noise / music
2. cut start / cut end      5. usable Ghanaian English
3. transcript mismatch      6. everything else

Don't reject a clip just because it isn't pristine. We are building a large
useful dataset, not studio recordings.
"""

RESULT_SHEETS = [
    ("GOOD", "KEEP"),
    ("BORDERLINE", "BORDERLINE"),
    ("BAD", "REJECT"),
]


def main() -> None:
    df = pd.read_csv(SHEET_CSV)
    n = len(df)
    last = n + 1  # last data row (header in row 1)

    wb = Workbook()

    # ---------- INSTRUCTIONS ----------
    ws = wb.active
    ws.title = "INSTRUCTIONS"
    ws.column_dimensions["A"].width = 110
    for i, line in enumerate(INSTRUCTIONS.splitlines(), start=1):
        c = ws.cell(row=i, column=1, value=line)
        if line and not line.startswith(" "):
            c.font = Font(bold=True)
        c.alignment = Alignment(wrap_text=False)

    # ---------- AUDIT ----------
    ws = wb.create_sheet("AUDIT")
    headers = ["listen_order", "wav_file", "stratum", "duration_s",
               "auto_flags", "corrected_text", "VERDICT", "REASON", "NOTES"]
    widths = [12, 52, 12, 11, 22, 90, 13, 22, 40]
    for j, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="2C343F")
        cell.font = Font(bold=True, color="DCE3EC")
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"

    for i, (_, r) in enumerate(df.iterrows(), start=2):
        ws.cell(row=i, column=1, value=int(r["listen_order"]))
        c = ws.cell(row=i, column=2, value=r["wav_file"])
        c.hyperlink = f"..\\data\\samples\\{r['wav_file']}"
        c.font = Font(color="0563C1", underline="single")
        ws.cell(row=i, column=3, value=r["stratum"])
        ws.cell(row=i, column=4, value=float(r["duration_ss"]))
        ws.cell(row=i, column=5, value=r["auto_flags"] if pd.notna(r["auto_flags"]) else "")
        tc = ws.cell(row=i, column=6, value=r["corrected_text"])
        tc.alignment = Alignment(wrap_text=True, vertical="top")

    # dropdowns
    dv_verdict = DataValidation(type="list", formula1='"' + ",".join(VERDICTS) + '"', allow_blank=True)
    dv_reason = DataValidation(type="list", formula1='"' + ",".join(REASONS) + '"', allow_blank=True)
    ws.add_data_validation(dv_verdict)
    ws.add_data_validation(dv_reason)
    dv_verdict.add(f"G2:G{last}")
    dv_reason.add(f"H2:H{last}")

    # colour coding
    ws.conditional_formatting.add(f"G2:G{last}", CellIsRule(
        operator="equal", formula=['"KEEP"'], fill=PatternFill("solid", fgColor="C6EFCE")))
    ws.conditional_formatting.add(f"G2:G{last}", CellIsRule(
        operator="equal", formula=['"BORDERLINE"'], fill=PatternFill("solid", fgColor="FFEB9C")))
    ws.conditional_formatting.add(f"G2:G{last}", CellIsRule(
        operator="equal", formula=['"REJECT"'], fill=PatternFill("solid", fgColor="FFC7CE")))

    # ---------- GOOD / BORDERLINE / BAD (live FILTER views) ----------
    for name, verdict in RESULT_SHEETS:
        s = wb.create_sheet(name)
        for j, h in enumerate(headers, start=1):
            cell = s.cell(row=1, column=j, value=h)
            cell.font = Font(bold=True)
            s.column_dimensions[get_column_letter(j)].width = widths[j - 1]
        formula = (f'=IFERROR(FILTER(AUDIT!A2:I{last},'
                   f'AUDIT!G2:G{last}="{verdict}"),'
                   f'"no {verdict} rows yet (needs Excel 365/2021)")')
        s.cell(row=2, column=1, value=formula)
        s.freeze_panes = "A2"

    wb.save(OUT)
    print(f"[done] {n} clips -> {OUT}")
    print("       sheets: INSTRUCTIONS, AUDIT, GOOD, BORDERLINE, BAD")


if __name__ == "__main__":
    sys.exit(main())
