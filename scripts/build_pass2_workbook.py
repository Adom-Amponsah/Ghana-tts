"""Build the pass-2 truncation-audit workbook (48 random 12-15 s clips).

Output: reports/pass2_truncation_audit.xlsx
Tabs:   INSTRUCTIONS, AUDIT (dropdowns + colour coding), GOOD, BAD.

Focus question: at the >=12 s floor, how often does a clip end (or start)
mid-sentence because of the 15 s hard ceiling?
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
MANIFEST = ROOT / "data" / "manifests" / "sample_pass2.csv"
OUT = ROOT / "reports" / "pass2_truncation_audit.xlsx"

VERDICTS = ["KEEP", "BORDERLINE", "REJECT"]
REASONS = ["cut_end", "cut_start", "cut_both", "multiple_speakers", "noise",
           "music", "transcript_mismatch", "unnatural_delivery", "other"]

INSTRUCTIONS = """PASS 2 — TRUNCATION CHECK (random 12-15 s clips)
=================================================

Purpose: confirm the >=12 s duration floor before the full 131 GB download.
48 clips, uniformly random shards + rows, all from the 12-15 s band, from 8
shards NOT used in the 200-clip audit.

THE MAIN QUESTION: does the clip end mid-sentence?
--------------------------------------------------
The corpus has a hard 15.000 s ceiling. Clips with at_ceiling = TRUE
(duration >= 14.9 s) are the prime suspects:

  BAD   "...the government would respond to the"   CUT
  GOOD  "...the government would respond immediately."  natural end

How to listen
-------------
1. Play the clip. Pay attention to the LAST 1-2 seconds first.
2. Then the FIRST second (does it start mid-word / mid-sentence?).
3. Read corrected_text afterwards - does it match what was said?

Marking
-------
VERDICT  KEEP / BORDERLINE / REJECT   (same meaning as pass 1)
REASON   cut_end / cut_start / cut_both / multiple_speakers / noise /
         music / transcript_mismatch / unnatural_delivery / other

What we need to know
--------------------
- reject rate among at_ceiling clips  (is the 15 s cut audible & damaging?)
- reject rate overall in the 12-15 s band (should be ~8% if pass 1 holds)

If at-ceiling clips keep at a similar rate, the >=12 s floor is confirmed
and we proceed to the full download.
"""


def main() -> None:
    df = pd.read_csv(MANIFEST)
    if "selection" not in df.columns:
        df["selection"] = "uniform_random"
    # ceiling-targeted suspects first, then uniform-random in seeded order
    ceiling = df[df["selection"] == "ceiling_target"].sort_values(
        "duration_ss", ascending=False)
    rest = df[df["selection"] != "ceiling_target"].sample(frac=1.0, random_state=777)
    df = pd.concat([ceiling, rest]).reset_index(drop=True)
    df["listen_order"] = df.index + 1
    n = len(df)
    last = n + 1

    wb = Workbook()

    ws = wb.active
    ws.title = "INSTRUCTIONS"
    ws.column_dimensions["A"].width = 100
    for i, line in enumerate(INSTRUCTIONS.splitlines(), start=1):
        c = ws.cell(row=i, column=1, value=line)
        if line and not line.startswith(" "):
            c.font = Font(bold=True)

    ws = wb.create_sheet("AUDIT")
    headers = ["listen_order", "wav_file", "shard", "duration_s", "selection",
               "corrected_text", "VERDICT", "REASON", "NOTES"]
    widths = [12, 48, 34, 11, 16, 90, 13, 22, 40]
    for j, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = PatternFill("solid", fgColor="2C343F")
        cell.font = Font(bold=True, color="DCE3EC")
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"

    for i, (_, r) in enumerate(df.iterrows(), start=2):
        ws.cell(row=i, column=1, value=int(r["listen_order"]))
        c = ws.cell(row=i, column=2, value=r["wav_file"])
        c.hyperlink = f"..\\data\\samples_pass2\\{r['wav_file']}"
        c.font = Font(color="0563C1", underline="single")
        ws.cell(row=i, column=3, value=r["shard"])
        ws.cell(row=i, column=4, value=float(r["duration_ss"]))
        cc = ws.cell(row=i, column=5, value=r["selection"])
        if r["selection"] == "ceiling_target":
            cc.font = Font(bold=True, color="B02418")
        tc = ws.cell(row=i, column=6, value=r["corrected_text"])
        tc.alignment = Alignment(wrap_text=True, vertical="top")

    dv_v = DataValidation(type="list", formula1='"' + ",".join(VERDICTS) + '"', allow_blank=True)
    dv_r = DataValidation(type="list", formula1='"' + ",".join(REASONS) + '"', allow_blank=True)
    ws.add_data_validation(dv_v)
    ws.add_data_validation(dv_r)
    dv_v.add(f"G2:G{last}")
    dv_r.add(f"H2:H{last}")
    ws.conditional_formatting.add(f"G2:G{last}", CellIsRule(
        operator="equal", formula=['"KEEP"'], fill=PatternFill("solid", fgColor="C6EFCE")))
    ws.conditional_formatting.add(f"G2:G{last}", CellIsRule(
        operator="equal", formula=['"BORDERLINE"'], fill=PatternFill("solid", fgColor="FFEB9C")))
    ws.conditional_formatting.add(f"G2:G{last}", CellIsRule(
        operator="equal", formula=['"REJECT"'], fill=PatternFill("solid", fgColor="FFC7CE")))

    for name, verdict in [("GOOD", "KEEP"), ("BAD", "REJECT")]:
        s = wb.create_sheet(name)
        for j, h in enumerate(headers, start=1):
            cell = s.cell(row=1, column=j, value=h)
            cell.font = Font(bold=True)
            s.column_dimensions[get_column_letter(j)].width = widths[j - 1]
        s.cell(row=2, column=1, value=(
            f'=IFERROR(FILTER(AUDIT!A2:I{last},AUDIT!G2:G{last}="{verdict}"),'
            f'"no {verdict} rows yet (needs Excel 365/2021)")'))
        s.freeze_panes = "A2"

    wb.save(OUT)
    n_ceiling = int((df["selection"] == "ceiling_target").sum())
    print(f"[done] {n} clips ({n_ceiling} ceiling-targeted) -> {OUT}")


if __name__ == "__main__":
    sys.exit(main())
