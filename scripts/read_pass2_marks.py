"""Read pass-2 truncation-audit verdicts from reports/pass2_truncation_audit.xlsx.

Splits results by selection block (ceiling_target vs uniform_random) and
prints verdict/reason breakdowns. Read-only on the workbook.
"""

import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "reports" / "pass2_truncation_audit.xlsx"
VERDICTS = ["KEEP", "BORDERLINE", "REJECT"]


def main() -> None:
    wb = load_workbook(XLSX, read_only=True)
    ws = wb["AUDIT"]

    blocks = {"ceiling_target": Counter(), "uniform_random": Counter()}
    reasons = Counter()
    notes = []
    total = Counter()

    for row in ws.iter_rows(min_row=2, values_only=True):
        order, wav, shard, dur, sel, text, verdict, reason, note = row[:9]
        if order is None:
            break
        sel = sel or "uniform_random"
        total[sel] += 1
        if not verdict:
            continue
        v = str(verdict).strip().upper()
        blocks[sel][v] += 1
        if reason:
            reasons[f"{sel}:{str(reason).strip().lower()}"] += 1
        if note:
            notes.append((order, v, str(note)))

    for sel, counts in blocks.items():
        n = sum(counts.values())
        print(f"{sel}  ({n}/{total[sel]} judged)")
        for v in VERDICTS:
            c = counts.get(v, 0)
            pct = 100 * c / n if n else 0
            print(f"  {v:<11} {c:>3}  ({pct:.0f}%)")
        print()

    if reasons:
        print("reasons:")
        for r, n in reasons.most_common():
            print(f"  {r:<40} {n}")
    if notes:
        print(f"\nnotes ({len(notes)}):")
        for order, v, note in notes:
            print(f"  #{order} [{v}] {note}")
    wb.close()


if __name__ == "__main__":
    sys.exit(main())
