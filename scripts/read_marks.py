"""Read listening verdicts filled into reports/listening_audit.xlsx.

Prints progress, verdict counts, reason breakdown, and notes. Read-only:
never writes to the workbook.
"""

import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "reports" / "listening_audit.xlsx"

VERDICTS = ["KEEP", "BORDERLINE", "REJECT"]


def main() -> None:
    wb = load_workbook(XLSX, read_only=True)
    ws = wb["AUDIT"]

    verdicts = Counter()
    reasons = Counter()
    flagged_verdicts = Counter()
    clean_verdicts = Counter()
    flagged_reasons = Counter()
    notes = []
    total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        order, wav, stratum, dur, auto_flags, text, verdict, reason, note = row[:9]
        if order is None:
            break
        total += 1
        if not verdict:
            continue
        v = str(verdict).strip().upper()
        verdicts[v] += 1
        flagged = bool(auto_flags)
        (flagged_verdicts if flagged else clean_verdicts)[v] += 1
        if reason:
            reasons[str(reason).strip().lower()] += 1
            if flagged:
                flagged_reasons[str(reason).strip().lower()] += 1
        if note:
            notes.append((order, wav, v, str(note)))

    judged = sum(verdicts.values())
    print(f"progress: {judged}/{total} judged")
    print("\nverdicts:")
    for v in VERDICTS:
        n = verdicts.get(v, 0)
        pct = 100 * n / judged if judged else 0
        print(f"  {v:<11} {n:>4}  ({pct:.1f}%)")

    print("\nauto-flagged suspects (DSP) vs your ears:")
    flagged_n = sum(flagged_verdicts.values())
    clean_n = sum(clean_verdicts.values())
    for v in VERDICTS:
        print(f"  {v:<11} flagged={flagged_verdicts.get(v,0):>3}/{flagged_n}   clean={clean_verdicts.get(v,0):>3}/{clean_n}")

    if reasons:
        print("\nreasons:")
        for r, n in reasons.most_common():
            print(f"  {r:<22} {n}")

    if notes:
        print(f"\nnotes ({len(notes)}):")
        for order, wav, v, note in notes:
            print(f"  #{order} [{v}] {note}")

    wb.close()


if __name__ == "__main__":
    sys.exit(main())
